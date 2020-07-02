# <pep8-80 compliant>

# import collections
# import json
import os
import sys

import openpyxl

import bpy

sys.stdout = sys.stderr


def update_proper_mat(io_files_path, skp_file):

    _su_file_name = skp_file.replace(".skp", "")

    all_skp_mats = []

    for skp_mat in bpy.data.materials:
        all_skp_mats.append(skp_mat.name)

    # with open(json_file) as json_data:
    #     j_data = json.load(json_data)

    source_files = os.listdir(io_files_path)

    for file in source_files:
        if file.endswith(".xlsx"):
            source_fpath = io_files_path + "\\" + file
            workbook = openpyxl.load_workbook(source_fpath)
            # sheets_name_list = workbook.get_sheet_names()
            sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])

            column_headers = []

            for i in range(1, sheet.max_column + 1):
                column_headers.append(sheet.cell(row=1, column=i).value)

            # print("\n")
            # print(column_headers)  # returns List
            # for value in sheet.iter_rows(min_row=1,
            #                              max_row=1,
            #                              values_only=True):
            #     print(value)  # returns Tuple
            # print("\n")

            shaders_dict = {}
            shader_d = {}

            for value in sheet.iter_rows(min_row=2,
                                         max_row=(sheet.max_row - 1),
                                         values_only=True):
                shaders_dict_id = "Shaders"

                # if value[0] == _su_file_name:
                #     shader_d.update({value[1]: value[2]})

                if (value[0] != None and value[1] != None):
                    shader_d.update({value[0]: value[1]})

                shaders_dict[shaders_dict_id] = shader_d

            # shaders_dict_json = json.dumps(shaders_dict)
            # shaders_dict_prime = json.loads(shaders_dict_json)

            # print("\n")
            # print(json.dumps(shaders_dict, indent=4, sort_keys=True))

    only_skp_mats_needed = []
    mapped_shaders_types = []

    # for key, value in j_data["Shaders"].items():
    #     if key in all_skp_mats:
    #         only_skp_mats_needed.append(key)
    #         mapped_shaders_types.append(value)

    for key, value in shaders_dict["Shaders"].items():
        for mat in all_skp_mats:
            if key in mat:
                only_skp_mats_needed.append(mat)
                mapped_shaders_types.append(value)

    # print("\n")
    # print(f"Material IDs Available In SketchUp File : {len(all_skp_mats)}")
    # print("*****")
    # for skp_mat in all_skp_mats:
    #     print(skp_mat)
    # print("\n")
    # u_only_skp_mats_needed = list(set(only_skp_mats_needed))
    # print(
    #     f"Material IDs Found In Spreadsheet Data  : {len(u_only_skp_mats_needed)}")
    # print("*****")
    # for skp_mat_xcl in u_only_skp_mats_needed:
    #     print(skp_mat_xcl)
    # print("\n")

    if bpy.data.objects.get("shaders_mesh") is None:

        mat_lib_path = bpy.path.abspath("//") + "mat_lib\\"
        source_file = os.listdir(mat_lib_path)[0]
        source = mat_lib_path + source_file

        data_libs = bpy.data.libraries
        with data_libs.load(source, link=True) as (data_from, data_to):
            data_to.objects = [name for name in data_from.objects]

        for obj in data_to.objects:
            if obj is not None:
                bpy.context.collection.objects.link(obj)

    for i in range(len(only_skp_mats_needed)):

        specific_mat = bpy.data.materials[only_skp_mats_needed[i]]
        specific_mat.use_nodes = True
        working_node_tree = specific_mat.node_tree
        mat_nodes = working_node_tree.nodes

        if mat_nodes.get('Principled BSDF'):

            target_node = mat_nodes['Principled BSDF']
            output_node = mat_nodes['Material Output']

            try:
                shader_group_req = bpy.data.node_groups[mapped_shaders_types[i]]

            except KeyError as e:
                print("Exception Raised: Invalid shader key is passed!")
                print("Error >> " + str(e))
                print("\n")

            except Exception as e:
                print(type(e), e)

            else:
                mat_nodes.new(type='ShaderNodeGroup'
                              ).node_tree = shader_group_req

                to_be_node = mat_nodes['Group']

                dif_col_input = target_node.inputs['Base Color']

                input_connected = not not dif_col_input.links

                if input_connected:
                    tex_node_col_socket = dif_col_input.links[0].from_socket
                else:
                    color_value = dif_col_input.default_value

                make_link = working_node_tree.links

                if input_connected:
                    make_link.new(tex_node_col_socket,
                                  to_be_node.inputs['Base Color'])
                else:
                    to_be_node.inputs['Base Color'].default_value = color_value

                make_link.new(output_node.inputs['Surface'],
                              to_be_node.outputs['Surface'])

                for output_socket in to_be_node.outputs:
                    if output_socket.name == 'Displacement':
                        make_link.new(output_node.inputs['Displacement'],
                                      to_be_node.outputs['Displacement'])

                mat_nodes.remove(target_node)
